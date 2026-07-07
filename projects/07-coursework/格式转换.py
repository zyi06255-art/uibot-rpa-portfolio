import os, json, copy, decimal, random
from apa_runtime import *
import openpyxl

def main(argument):
    """
    对所有分析Excel中的数值设置百分比/数字格式
    文件列表：
      1. 药明康德盈利能力分析.xlsx
      2. 药明康德偿债能力分析.xlsx
      3. 药明康德综合能力分析.xlsx
    """
    res_path = r"D:\laiye project\课程作业\res"
    stock_name = "药明康德"

    files = [
        stock_name + "盈利能力分析.xlsx",
        stock_name + "偿债能力分析.xlsx",
        stock_name + "综合能力分析.xlsx",
    ]

    for filename in files:
        filepath = os.path.join(res_path, filename)
        if not os.path.exists(filepath):
            print("⚠ 跳过（不存在）: {}".format(filepath))
            continue

        print("处理: {}".format(filename))
        wb = openpyxl.load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 找到数据区域：从第4或第5行开始（跳过标题），B/C列是数值
            for row in range(4, ws.max_row + 1):
                for col in [2, 3]:  # B列(2023年) 和 C列(2024年)
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        # 判断是否为数字
                        try:
                            val = float(cell.value)
                            # 已是百分比形式（如45.3=45.3%），用0.00格式
                            # 如果是小于1的小数(原始比率)，用0.00%格式
                            if abs(val) < 1 and val != 0:
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = '0.00'
                        except (ValueError, TypeError):
                            pass

                # D列(2025年)同样处理
                cell_d = ws.cell(row=row, column=4)
                if cell_d.value is not None:
                    try:
                        val = float(cell_d.value)
                        if abs(val) < 1 and val != 0:
                            cell_d.number_format = '0.00%'
                        else:
                            cell_d.number_format = '0.00'
                    except (ValueError, TypeError):
                        pass

            # 调整列宽
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

        wb.save(filepath)
        wb.close()
        print("  ✓ 格式化完成")

    print("========== 格式转换全部完成 ==========")
    return {"status": "ok"}
