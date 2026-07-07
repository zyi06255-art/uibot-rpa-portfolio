import os, json, copy, decimal, random
from apa_runtime import *
import openpyxl

def main(argument):
    res_path = GlobalVariables["$Flow.ResPath"]
    file_path = os.path.join(res_path, "财务报表与财务分析.xlsx")

    wb = openpyxl.load_workbook(file_path)

    sheets = ["盈利能力分析", "偿债能力分析"]

    for sn in sheets:
        if sn in wb.sheetnames:
            ws = wb[sn]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=4)
                if cell.value is not None:
                    cell.number_format = '0.00%'
            for row in range(2, ws.max_row + 1):
                for col in [2, 3]:
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.number_format = '0.00%'

    wb.save(file_path)
    wb.close()

    return {"status": "ok"}
